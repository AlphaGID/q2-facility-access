"""
Step 2: Reconcile LGA_SEN_Districts.xlsx into a clean, normalized crosswalk.

The source workbook (data/raw/LGA_SEN_Districts.xlsx) has:
  - Sheet1: the current crosswalk. Merged title/header cells; State and
    Senatorial District are only written on the first row of each block
    (blank cells below mean "same as above"). A free-text Remarks column
    sometimes carries substantive information (renames, transfers,
    contested boundaries).
  - old_2019: an explicitly superseded prior version ("SUPERSEDED 2019
    LIST. DO NOT USE." in cell A1). Read for the record, never used.

Output:
  - data/processed/lga_sen_crosswalk.csv: one row per lga_code, ready to
    load into the spatial database.
  - outputs/logs/crosswalk_reconciliation_log.md: human-readable account
    of every row that needed a decision, and what was decided.
"""
import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent))
import config

import openpyxl
import pandas as pd
import sqlite3


def normalize_lga_name(name):
    if name is None:
        return None
    s = str(name).strip().upper()
    s = s.replace(" LOCAL GOVT AREA", "").replace(" LGA", "")
    s = s.replace("-", "").replace(" ", "")
    return s


def load_sheet1_raw():
    wb = openpyxl.load_workbook(config.F_LGA_SEN_XLSX, data_only=True)
    ws = wb["Sheet1"]
    rows = list(ws.iter_rows(min_row=7, max_row=ws.max_row, values_only=True))
    return rows


def load_old_2019_note():
    wb = openpyxl.load_workbook(config.F_LGA_SEN_XLSX, data_only=True)
    ws = wb["old_2019"]
    a1 = ws["A1"].value
    return a1, ws.max_row


def forward_fill_parse(rows):
    """Turn raw sheet rows into records, forward-filling state and
    senatorial district, and dropping fully blank spacer rows."""
    records = []
    state = None
    sen = None
    for r in rows:
        if len(r) < 6:
            continue
        st, lga, send, nwards, code, remarks = r[:6]
        if st:
            state = str(st).strip()
        if send:
            sen = str(send).strip()
        if lga is None and code is None:
            continue
        records.append({
            "state_name": state,
            "lga_name_raw": str(lga).strip() if lga else None,
            "lga_name_norm": normalize_lga_name(lga),
            "sen_district": sen,
            "n_wards": nwards,
            "lga_code": str(code).strip() if code else None,
            "remarks": str(remarks).strip() if remarks else None,
        })
    return pd.DataFrame(records)


def load_gpkg_lgas():
    con = sqlite3.connect(config.F_ADMIN_BOUNDARIES_GPKG)
    df = pd.read_sql("SELECT lga_code, lga_name, sen_code, sen_district, state_name FROM lgas", con)
    con.close()
    return df


def reconcile():
    log_lines = ["# LGA to Senatorial District Crosswalk — Reconciliation Log\n"]

    note, n = load_old_2019_note()
    log_lines.append(f"- `old_2019` sheet found ({n} rows). Header cell A1 reads: "
                      f"\"{note}\". Excluded from processing as instructed by the source file itself.\n")

    raw = load_sheet1_raw()
    df = forward_fill_parse(raw)
    log_lines.append(f"- Parsed {len(df)} data rows from `Sheet1`, after forward-filling State and "
                      f"Senatorial District and dropping blank spacer rows.\n")

    dup_mask = df.duplicated(subset=["lga_code"], keep=False)
    dups = df[dup_mask].copy()

    exact_dupes, conflicts = [], []
    for code, grp in dups.groupby("lga_code"):
        grp_check = grp.drop(columns=["lga_code"])
        if grp_check.drop_duplicates().shape[0] == 1:
            exact_dupes.append(code)
        else:
            conflicts.append(code)

    log_lines.append(f"\n## Duplicate LGA codes\n")
    log_lines.append(f"- {len(exact_dupes)} code(s) appeared as exact repeated rows: {exact_dupes}. "
                      f"Treated as accidental duplication in data entry; extra copies dropped.\n")
    log_lines.append(f"- {len(conflicts)} code(s) appeared with genuinely conflicting senatorial "
                      f"district assignments: {conflicts}.\n")

    gpkg = load_gpkg_lgas()

    resolved_rows, conflict_detail = [], []
    for code in conflicts:
        grp = dups[dups.lga_code == code]
        gpkg_row = gpkg[gpkg.lga_code == code]
        gpkg_sen = gpkg_row.sen_district.iloc[0] if len(gpkg_row) else None
        chosen = None
        for _, r in grp.iterrows():
            if r.sen_district == gpkg_sen:
                chosen = r
        detail = (f"  - `{code}`: xlsx gives " +
                  " vs ".join(f'"{r.sen_district}" (remarks: {r.remarks})' for _, r in grp.iterrows()) +
                  f'. `admin_boundaries.gpkg` (current authoritative boundary layer) has sen_district='
                  f'"{gpkg_sen}". Resolved to the boundary-layer value; the other xlsx row is treated '
                  f'as superseded/not reflected in the current boundary and flagged here, not deleted.')
        conflict_detail.append(detail)
        if chosen is not None:
            resolved_rows.append(chosen)
        else:
            resolved_rows.append(grp.iloc[0])
            conflict_detail.append(f"    WARNING: neither xlsx variant for {code} matched the "
                                    f"boundary layer. Kept the first row; needs human review.")

    if conflict_detail:
        log_lines.append("\n".join(conflict_detail) + "\n")

    final = df[~df.lga_code.isin(dups.lga_code.unique())].copy()
    for code in exact_dupes:
        final = pd.concat([final, dups[dups.lga_code == code].iloc[[0]]], ignore_index=True)
    if resolved_rows:
        final = pd.concat([final, pd.DataFrame(resolved_rows)], ignore_index=True)

    assert final.lga_code.is_unique, "final crosswalk should have one row per lga_code"

    xlsx_codes = set(final.lga_code)
    gpkg_codes = set(gpkg.lga_code)
    missing_in_xlsx = gpkg_codes - xlsx_codes
    extra_in_xlsx = xlsx_codes - gpkg_codes

    log_lines.append(f"\n## Coverage check against `admin_boundaries.gpkg` ({len(gpkg_codes)} LGAs)\n")
    log_lines.append(f"- In boundary layer but missing from crosswalk: {len(missing_in_xlsx)} "
                      f"{sorted(missing_in_xlsx) if missing_in_xlsx else ''}\n")
    log_lines.append(f"- In crosswalk but not in boundary layer: {len(extra_in_xlsx)} "
                      f"{sorted(extra_in_xlsx) if extra_in_xlsx else ''}\n")
    log_lines.append(f"- Final reconciled crosswalk: {len(final)} rows, one per LGA code.\n")

    final = final.sort_values("lga_code").reset_index(drop=True)
    final.to_csv(config.PROCESSED / "lga_sen_crosswalk.csv", index=False)

    log_path = config.LOGS / "crosswalk_reconciliation_log.md"
    log_path.write_text("\n".join(log_lines))

    print(f"Wrote {config.PROCESSED / 'lga_sen_crosswalk.csv'} ({len(final)} rows)")
    print(f"Wrote {log_path}")
    print()
    print(final.head(10).to_string())
    return final


if __name__ == "__main__":
    reconcile()
